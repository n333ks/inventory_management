#!/usr/bin/env python3
"""
Recalculate In-Stock, In-Production, Pre-Sale QTY and Variance for every
variant in inventory_master.xlsx, writing static values instead of formulas.

Replaces COUNTIFS/formula approach which breaks when openpyxl inserts rows
(openpyxl does not update formula cell references on insert_rows).

Summary row = first row encountered for each (Design Name, Size, Finish,
Swing, Glass Type) combination.  Counts + Variance are written there.
All other rows in the block have those columns cleared.
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

INVENTORY_FILE = "inventory_master.xlsx"
DATA_START = 3
VARIANT_COLS = range(1, 6)   # A–E

C_INSTOCK  = 6    # F
C_INPROD   = 7    # G
C_PRESALE  = 8    # H
C_OPTIMAL  = 9    # I
C_VARIANCE = 10   # J

CENTER = Alignment(horizontal="center")
FONT   = Font(name="Arial", size=10, color="FF000000")
FONT_BOLD = Font(name="Arial", size=10, bold=True, color="FF000000")
FONT_RED  = Font(name="Arial", size=10, color="FFFF0000")


def _variant_key(ws, r):
    return tuple(ws.cell(r, c).value for c in VARIANT_COLS)


def _status_bucket(status):
    """Map a status string to the count bucket it belongs to."""
    if not status:
        return None
    s = str(status)
    if s == "In Stock":
        return "in_stock"
    if s.startswith("Pre-Sale"):
        return "pre_sale"
    if s == "In Production":
        return "in_prod"
    return None


def refresh_counts(ws):
    """
    Scan ws, tally status counts per variant, write static values to the
    summary (first) row of each variant.  Returns list of summary row numbers.
    """
    last_row = ws.max_row

    # ── Pass 1: collect counts and first/all rows per variant ─────────────────
    counts   = {}   # variant_key → {in_stock, in_prod, pre_sale}
    first_row = {}  # variant_key → row number of summary row
    all_rows  = {}  # variant_key → [all row numbers]

    for r in range(DATA_START, last_row + 1):
        key = _variant_key(ws, r)
        if not any(key):
            continue
        if key not in first_row:
            first_row[key] = r
            counts[key]    = {"in_stock": 0, "in_prod": 0, "pre_sale": 0}
            all_rows[key]  = []
        all_rows[key].append(r)

        bucket = _status_bucket(ws.cell(r, 12).value)  # col L = Status
        if bucket:
            counts[key][bucket] += 1

    # ── Pass 2: write to summary rows, clear detail rows ──────────────────────
    summary_rows = []
    for key, sr in first_row.items():
        c = counts[key]
        in_stock = c["in_stock"]
        in_prod  = c["in_prod"]
        pre_sale = c["pre_sale"]
        optimal  = ws.cell(sr, C_OPTIMAL).value or 0
        variance = in_stock + in_prod + pre_sale - optimal

        for col, val in [
            (C_INSTOCK,  in_stock),
            (C_INPROD,   in_prod),
            (C_PRESALE,  pre_sale),
            (C_VARIANCE, variance),
        ]:
            cell = ws.cell(sr, col)
            cell.value     = val
            cell.alignment = CENTER
            cell.font      = FONT_RED if (col == C_VARIANCE and variance < 0) else FONT

        summary_rows.append(sr)

        # Clear count/variance on every other row in this variant's block
        for r in all_rows[key]:
            if r == sr:
                continue
            for col in (C_INSTOCK, C_INPROD, C_PRESALE, C_VARIANCE):
                cell = ws.cell(r, col)
                cell.value = None

    return summary_rows


def main():
    wb = load_workbook(INVENTORY_FILE)
    ws  = wb.active
    summary_rows = refresh_counts(ws)
    wb.save(INVENTORY_FILE)
    print(f"Refreshed counts for {len(summary_rows)} variant(s) → {INVENTORY_FILE}")


if __name__ == "__main__":
    main()
