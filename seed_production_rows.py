#!/usr/bin/env python3
"""
Seed In Production rows for variants that appear in the upcoming manifest.

For each target variant + quantity:
  1. Removes any ghost rows for that variant (variant data present but no serial/status)
  2. Inserts the specified number of In Production rows after the last existing row
     of that variant block (blank serial, Status = "In Production")

Run this before parse_manifest.py to simulate units being ordered and put into production.
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from apply_status_fills import apply_fills
from fix_inventory_formatting import apply_design_block_borders
from refresh_counts import refresh_counts

INVENTORY_FILE = "inventory_master.xlsx"
C_SERIAL = 11   # K
C_STATUS = 12   # L
DATA_START = 3

WHITE     = "FFFFFFFF"
BLACK     = "FF000000"
NO_FILL   = PatternFill(fill_type=None)

# Variants to seed and how many In Production rows to add.
# (Design Name, Size, Finish, Swing, Glass Type): qty
SEED = {
    ("Valencia Single Door", '32" x 80"', "Matte Black", "Left Inswing", "Clear"): 3,
    ("Sevilla French Door",  '36" x 80"', "Satin White", "Left Inswing", "None"):  2,
    ("Altea Double Door",    '36" x 80"', "Matte Black", "Left Inswing", "Rain"):  2,
}


def row_variant(ws, r):
    return tuple(ws.cell(r, c).value for c in range(1, 6))


def is_ghost(ws, r):
    """Row has variant data but no serial and no status — leftover from test/manual edit."""
    has_variant = any(ws.cell(r, c).value for c in range(1, 6))
    no_serial   = not ws.cell(r, C_SERIAL).value
    no_status   = not ws.cell(r, C_STATUS).value
    return has_variant and no_serial and no_status


def write_production_row(ws, r, variant):
    design, size, finish, swing, glass = variant
    values = [design, size, finish, swing, glass]
    for c, val in enumerate(values, start=1):
        cell = ws.cell(r, c)
        cell.value = val
        cell.font  = Font(name="Arial", size=10, color=WHITE)
        cell.alignment = Alignment(horizontal="left")
    # Serial blank, Status = In Production (centered)
    ws.cell(r, C_SERIAL).value = None
    status_cell = ws.cell(r, C_STATUS)
    status_cell.value = "In Production"
    status_cell.font  = Font(name="Arial", size=10, color=BLACK)
    status_cell.alignment = Alignment(horizontal="center")


def main():
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active

    for variant, qty in SEED.items():
        # Find all rows belonging to this variant
        variant_rows = [r for r in range(DATA_START, ws.max_row + 1)
                        if row_variant(ws, r) == variant]

        # Remove ghost rows (bottom-to-top so indices stay valid)
        ghosts = [r for r in variant_rows if is_ghost(ws, r)]
        for r in sorted(ghosts, reverse=True):
            ws.delete_rows(r)
            print(f"  Removed ghost row {r} for {variant[0]} {variant[1]}")

        # Re-scan after deletions
        variant_rows = [r for r in range(DATA_START, ws.max_row + 1)
                        if row_variant(ws, r) == variant]

        if not variant_rows:
            print(f"  WARNING: no existing rows found for {variant[0]} {variant[1]} — skipping")
            continue

        last_row = max(variant_rows)
        insert_at = last_row + 1

        ws.insert_rows(insert_at, amount=qty)
        for i in range(qty):
            write_production_row(ws, insert_at + i, variant)

        print(f"  Seeded {qty} In Production row(s) for {variant[0]} {variant[1]} {variant[2]} "
              f"after row {last_row}")

    apply_design_block_borders(ws)
    refresh_counts(ws)
    apply_fills(ws)
    wb.save(INVENTORY_FILE)
    print(f"\nSaved → {INVENTORY_FILE}")
    print("Open in Excel — new In Production rows should appear in red with blank serials.")


if __name__ == "__main__":
    main()
